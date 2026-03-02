"""
신구조문 대조표 .xlsx 생성.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


class XlsxComparisonWriter:
    """신구조문 대조표 생성기 (.xlsx)."""

    def generate(
        self,
        changes: list[dict],
        output_path: str,
        company_name: str = "",
    ) -> str:
        path = Path(output_path)
        path.parent.mkdir(parents=True, exist_ok=True)

        if path.suffix.lower() != ".xlsx":
            path = path.with_suffix(".xlsx")

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "신구조문 대조표"

        header_font = Font(name="맑은 고딕", bold=True, size=11)
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_font = Font(name="맑은 고딕", size=10)
        cell_align = Alignment(vertical="top", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        row = 1
        if company_name:
            worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
            title_cell = worksheet.cell(
                row=1, column=1, value=f"{company_name} 취업규칙 신구조문 대조표"
            )
            title_cell.font = Font(name="맑은 고딕", bold=True, size=14)
            title_cell.alignment = Alignment(horizontal="center")
            row = 3

        headers = ["조문번호", "현행 규정", "변경 규정", "변경 사유"]
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        worksheet.column_dimensions["A"].width = 12
        worksheet.column_dimensions["B"].width = 45
        worksheet.column_dimensions["C"].width = 45
        worksheet.column_dimensions["D"].width = 40

        for change in changes:
            row += 1
            article_num = change.get("article_number", "")
            values = [
                f"제{article_num}조" if article_num else "",
                change.get("old_text", ""),
                change.get("new_text", ""),
                change.get("reason", ""),
            ]

            for col, value in enumerate(values, 1):
                cell = worksheet.cell(row=row, column=col, value=value)
                cell.font = cell_font
                cell.alignment = cell_align
                cell.border = thin_border

        workbook.save(str(path))
        return str(path)
